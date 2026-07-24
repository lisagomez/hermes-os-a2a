"""Tests del departamento de Procesos — gates deterministas del Supervisor.

Estilo de los otros servicios del trío: nombres descriptivos en español, gates
binarios sin LLM, y la invariante clave: un gate que no se puede correr NO se
asume aprobado (no_ejecutable = rechazo).

Correr:  pytest -q supervisor-a2a/test_procesos.py
"""
import copy
from pathlib import Path

import pytest

import chequeos_procesos as chk

# --- Fixtures: un paquete to-be VÁLIDO como base -----------------------------

DIAG_BASE = {
    "proyecto": "comnorte-clasificador-facturas",
    "cliente": "Comercializadora del Norte",
    "alcance": "mediano",
    "supuestos": ["Volumen estimado 1200 facturas/mes (a confirmar)"],
    "linea_base": {
        "volumen_mes": 200, "horas_humano_corrida": 1.5,
        "costo_hora_op_usd": 25, "costo_actual_mensual_usd": 7500,
        "costo_actual_anual_usd": 90000, "es_estimado": True,
        "supuestos": ["Costo-hora operativo asumido en 25 USD (a validar)"],
    },
    "consejo": "Estandarizar la entrada y automatizar con revisión humana.",
    "reto_limitantes": [
        "Contpaqi sin API documentada: validar antes de cotizar.",
        "Contabilizar es irreversible: gate humano obligatorio.",
    ],
    "pasos_as_is": [
        {"id": 1, "paso": "Descargar el PDF del correo",
         "responsable": "Auxiliar", "sistema": "Gmail",
         "veredicto_esoa": "simplificar",
         "justificacion": "Handoff manual; se recibe en bandeja estándar."},
        {"id": 2, "paso": "Determinar si el gasto es deducible",
         "responsable": "Contador", "sistema": "Excel",
         "veredicto_esoa": "automatizar",
         "justificacion": "Decisión con reglas fiscales; agente + grafo."},
    ],
    "cinco_s": {
        "seiri_clasificar": "Se capturan 2 columnas que nadie usa.",
        "seiton_ordenar": "PDFs en correos personales; sin ubicación estándar.",
        "seiso_limpiar": "RFC falta en ~10% de los PDFs.",
        "seiketsu_estandarizar": "Cada proveedor manda formato distinto.",
        "shitsuke_disciplina": "Sin validación que sostenga el estándar.",
    },
    "diseno_a2a": [
        {"automatizacion": "Clasificación de deducibilidad",
         "complejidad": "media",
         "agentes": ["agente-clasificador-fiscal"],
         "coordinacion_a2a": "consulta grafo-a2a con fuente",
         "integraciones": ["Contpaqi API", "grafo"],
         "control_humano": "revisar las 'dudoso' antes de contabilizar",
         "manejo_error": "sin regla aplicable -> dudoso -> escala a humano"},
    ],
}

SPEC_BASE = {
    "proyecto": "comnorte-clasificador-facturas",
    "alcance": "mediano",
    "resumen": "Automatizar clasificación con revisión humana.",
    "stack_cliente": {"suite": "M365",
                      "herramientas": ["Power Automate", "Copilot", "Contpaqi"]},
    "construir": [
        {"id": "cls-01", "automatizacion": "Clasificación de deducibilidad",
         "complejidad": "media", "departamento_destino": "software",
         "sdd": True, "spec_ref": "specs/clasificador-facturas.md",
         "skills_requeridas": ["trio-software"],
         "clis_requeridos": ["contpaqi", "grafo"],
         "herramientas_propuestas": ["Power Automate", "Copilot"],
         "justificacion_herramientas": "",
         "integraciones": ["Contpaqi API"],
         "control_humano": "revisar dudosas antes de contabilizar",
         "gate_humano_irreversible": True},
    ],
    "disparo": {"cola": True, "requiere_aprobacion_humana": True,
                "tope_gasto_usd": 5, "orden": ["cls-01"]},
}

REPORTE_BASE = (
    "# Diagnóstico — Clasificación de facturas — Comercializadora del Norte\n"
    "## 1. Resumen ejecutivo\nProceso manual; agente + grafo con revisión.\n"
    "## 4. Análisis ESOA\nVeredictos por paso.\n"
    "## 8. ROI y tiempo humano-agente\nAhorro y payback. Fuente: grafo (LISR 27).\n"
)


def _xlsx_valido(destino: Path):
    """xlsx mínimo con los tokens que exige el gate (USD, MXN, tipo de cambio)."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Supuestos"
    ws.append(["Tipo de cambio (MXN por USD)", 18.5])
    ws.append(["Margen sobre costo", 0.35])
    ws2 = wb.create_sheet("Presupuesto")
    ws2.append(["Concepto", "USD", "MXN"])
    ws2.append(["PRECIO", 29270, 541497])
    wb.save(destino)


def _dump_yaml(obj) -> str:
    import yaml
    return yaml.safe_dump(obj, allow_unicode=True, sort_keys=False)


@pytest.fixture
def worktree(tmp_path):
    """Escribe un paquete to-be válido y devuelve la ruta."""
    (tmp_path / "diagnostico.yaml").write_text(_dump_yaml(DIAG_BASE), encoding="utf-8")
    (tmp_path / "build-spec.yaml").write_text(_dump_yaml(SPEC_BASE), encoding="utf-8")
    (tmp_path / "reporte.md").write_text(REPORTE_BASE, encoding="utf-8")
    _xlsx_valido(tmp_path / "presupuesto.xlsx")
    return tmp_path


def _reescribe(worktree: Path, nombre: str, obj):
    (worktree / nombre).write_text(_dump_yaml(obj), encoding="utf-8")


# --- Camino feliz: el paquete válido pasa TODOS los gates --------------------

def test_paquete_valido_aprueba_todos_los_gates(worktree):
    ok, resultados = chk.correr_todos(worktree)
    fallidos = [n for n, ch in resultados if not ch.passed]
    assert ok, f"gates fallidos inesperadamente: {fallidos}"


# --- estructura_diagnostico -------------------------------------------------

def test_estructura_rechaza_si_falta_un_artefacto(worktree):
    (worktree / "presupuesto.xlsx").unlink()
    ctx = chk._cargar(worktree)
    ch = chk.estructura_diagnostico(ctx)
    assert not ch.passed and "presupuesto.xlsx" in ch.hallazgo


def test_estructura_rechaza_reporte_sin_seccion(worktree):
    (worktree / "reporte.md").write_text("# sin secciones\n", encoding="utf-8")
    ctx = chk._cargar(worktree)
    assert not chk.estructura_diagnostico(ctx).passed


def test_estructura_rechaza_alcance_invalido(worktree):
    d = copy.deepcopy(DIAG_BASE); d["alcance"] = "enorme"
    _reescribe(worktree, "diagnostico.yaml", d)
    ctx = chk._cargar(worktree)
    assert not chk.estructura_diagnostico(ctx).passed


# --- linea_base_cuantificada ------------------------------------------------

def test_linea_base_no_ejecutable_si_falta(worktree):
    d = copy.deepcopy(DIAG_BASE); d.pop("linea_base")
    _reescribe(worktree, "diagnostico.yaml", d)
    ctx = chk._cargar(worktree)
    ch = chk.linea_base_cuantificada(ctx)
    assert not ch.passed and ch.no_ejecutable


def test_linea_base_rechaza_sin_costo(worktree):
    d = copy.deepcopy(DIAG_BASE)
    d["linea_base"]["costo_actual_mensual_usd"] = 0
    d["linea_base"]["costo_actual_anual_usd"] = 0
    _reescribe(worktree, "diagnostico.yaml", d)
    ctx = chk._cargar(worktree)
    assert not chk.linea_base_cuantificada(ctx).passed


def test_linea_base_estimada_exige_supuestos(worktree):
    d = copy.deepcopy(DIAG_BASE)
    d["linea_base"]["es_estimado"] = True
    d["linea_base"]["supuestos"] = []
    _reescribe(worktree, "diagnostico.yaml", d)
    ctx = chk._cargar(worktree)
    assert not chk.linea_base_cuantificada(ctx).passed


# --- consejo_y_reto ---------------------------------------------------------

def test_consejo_y_reto_rechaza_sin_consejo(worktree):
    d = copy.deepcopy(DIAG_BASE); d["consejo"] = ""
    _reescribe(worktree, "diagnostico.yaml", d)
    ctx = chk._cargar(worktree)
    assert not chk.consejo_y_reto(ctx).passed


def test_consejo_y_reto_rechaza_reto_vacio(worktree):
    d = copy.deepcopy(DIAG_BASE); d["reto_limitantes"] = []
    _reescribe(worktree, "diagnostico.yaml", d)
    ctx = chk._cargar(worktree)
    ch = chk.consejo_y_reto(ctx)
    assert not ch.passed and "reto" in ch.hallazgo.lower()


# --- herramientas_en_stack --------------------------------------------------

def test_stack_rechaza_herramienta_fuera_sin_justificacion(worktree):
    s = copy.deepcopy(SPEC_BASE)
    s["construir"][0]["herramientas_propuestas"] = ["Zapier"]  # fuera de M365
    s["construir"][0]["justificacion_herramientas"] = ""
    _reescribe(worktree, "build-spec.yaml", s)
    ctx = chk._cargar(worktree)
    ch = chk.herramientas_en_stack(ctx)
    assert not ch.passed and "Zapier" in ch.hallazgo


def test_stack_acepta_herramienta_fuera_con_justificacion(worktree):
    s = copy.deepcopy(SPEC_BASE)
    s["construir"][0]["herramientas_propuestas"] = ["Zapier"]
    s["construir"][0]["justificacion_herramientas"] = \
        "M365 no cubre este conector; Zapier es la vía más barata."
    _reescribe(worktree, "build-spec.yaml", s)
    ctx = chk._cargar(worktree)
    assert chk.herramientas_en_stack(ctx).passed


def test_stack_rechaza_propuesta_sin_declarar_stack(worktree):
    s = copy.deepcopy(SPEC_BASE); s.pop("stack_cliente")
    _reescribe(worktree, "build-spec.yaml", s)
    ctx = chk._cargar(worktree)
    assert not chk.herramientas_en_stack(ctx).passed


# --- esoa_completo ----------------------------------------------------------

def test_esoa_rechaza_veredicto_invalido(worktree):
    d = copy.deepcopy(DIAG_BASE)
    d["pasos_as_is"][1]["veredicto_esoa"] = "teletransportar"
    _reescribe(worktree, "diagnostico.yaml", d)
    ctx = chk._cargar(worktree)
    ch = chk.esoa_completo(ctx)
    assert not ch.passed and "veredicto_esoa" in ch.hallazgo


def test_esoa_rechaza_justificacion_vacia(worktree):
    d = copy.deepcopy(DIAG_BASE); d["pasos_as_is"][0]["justificacion"] = ""
    _reescribe(worktree, "diagnostico.yaml", d)
    ctx = chk._cargar(worktree)
    assert not chk.esoa_completo(ctx).passed


def test_esoa_no_ejecutable_si_no_hay_pasos(worktree):
    d = copy.deepcopy(DIAG_BASE); d.pop("pasos_as_is")
    _reescribe(worktree, "diagnostico.yaml", d)
    ctx = chk._cargar(worktree)
    ch = chk.esoa_completo(ctx)
    assert not ch.passed and ch.no_ejecutable


# --- cinco_s_aplicado -------------------------------------------------------

def test_cinco_s_rechaza_si_falta_una_ese(worktree):
    d = copy.deepcopy(DIAG_BASE); d["cinco_s"].pop("shitsuke_disciplina")
    _reescribe(worktree, "diagnostico.yaml", d)
    ctx = chk._cargar(worktree)
    ch = chk.cinco_s_aplicado(ctx)
    assert not ch.passed and "shitsuke" in ch.hallazgo


def test_cinco_s_acepta_na_con_razon(worktree):
    d = copy.deepcopy(DIAG_BASE)
    d["cinco_s"]["seiso_limpiar"] = "n/a: los datos ya llegan limpios del ERP."
    _reescribe(worktree, "diagnostico.yaml", d)
    ctx = chk._cargar(worktree)
    assert chk.cinco_s_aplicado(ctx).passed


# --- control_humano_por_automatizacion --------------------------------------

def test_control_humano_rechaza_si_falta(worktree):
    d = copy.deepcopy(DIAG_BASE); d["diseno_a2a"][0]["control_humano"] = ""
    _reescribe(worktree, "diagnostico.yaml", d)
    ctx = chk._cargar(worktree)
    assert not chk.control_humano_por_automatizacion(ctx).passed


def test_control_humano_rechaza_cero_humanos(worktree):
    d = copy.deepcopy(DIAG_BASE)
    d["diseno_a2a"][0]["control_humano"] = "cero humanos, totalmente autónomo"
    _reescribe(worktree, "diagnostico.yaml", d)
    ctx = chk._cargar(worktree)
    ch = chk.control_humano_por_automatizacion(ctx)
    assert not ch.passed and "cero humanos" in ch.hallazgo


# --- presupuesto_dos_monedas ------------------------------------------------

def test_presupuesto_rechaza_una_sola_moneda(worktree):
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active
    ws.append(["Tipo de cambio", 18.5]); ws.append(["PRECIO", "USD", 29270])
    wb.save(worktree / "presupuesto.xlsx")   # sin MXN
    ctx = chk._cargar(worktree)
    assert not chk.presupuesto_dos_monedas(ctx).passed


def test_presupuesto_rechaza_sin_tipo_de_cambio(worktree):
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active
    ws.append(["Concepto", "USD", "MXN"]); ws.append(["PRECIO", 1, 18])
    wb.save(worktree / "presupuesto.xlsx")   # sin TC/supuesto
    ctx = chk._cargar(worktree)
    assert not chk.presupuesto_dos_monedas(ctx).passed


# --- build_spec_valida (el candado de aprobación humana) --------------------

def test_build_spec_rechaza_sin_aprobacion_humana(worktree):
    s = copy.deepcopy(SPEC_BASE)
    s["disparo"]["requiere_aprobacion_humana"] = False
    _reescribe(worktree, "build-spec.yaml", s)
    ctx = chk._cargar(worktree)
    ch = chk.build_spec_valida(ctx)
    assert not ch.passed and "aprobacion_humana" in ch.hallazgo


def test_build_spec_rechaza_item_incompleto(worktree):
    s = copy.deepcopy(SPEC_BASE)
    s["construir"][0].pop("clis_requeridos")
    _reescribe(worktree, "build-spec.yaml", s)
    ctx = chk._cargar(worktree)
    assert not chk.build_spec_valida(ctx).passed


def test_build_spec_rechaza_gate_no_booleano(worktree):
    s = copy.deepcopy(SPEC_BASE)
    s["construir"][0]["gate_humano_irreversible"] = "sí"
    _reescribe(worktree, "build-spec.yaml", s)
    ctx = chk._cargar(worktree)
    assert not chk.build_spec_valida(ctx).passed


# --- sin_marcadores (marca blanca) ------------------------------------------

def test_marcadores_rechaza_placeholder_sin_sustituir(worktree):
    d = copy.deepcopy(DIAG_BASE); d["cliente"] = "[CLIENTE]"
    _reescribe(worktree, "diagnostico.yaml", d)
    ctx = chk._cargar(worktree)
    ch = chk.sin_marcadores(ctx)
    assert not ch.passed and "[CLIENTE]" in ch.hallazgo


# --- fuentes_citadas --------------------------------------------------------

def test_fuentes_rechaza_afirmacion_fiscal_sin_fuente(worktree):
    d = copy.deepcopy(DIAG_BASE)
    # afirmación fiscal, y quitamos toda señal de fuente/grafo del diseño y reporte
    d["diseno_a2a"][0]["coordinacion_a2a"] = "el agente decide deducible por su cuenta"
    d["diseno_a2a"][0]["integraciones"] = ["Contpaqi API"]
    d["pasos_as_is"][1]["justificacion"] = "Es deducible el gasto, sin más."
    _reescribe(worktree, "diagnostico.yaml", d)
    (worktree / "reporte.md").write_text(
        "## 1. Resumen ejecutivo\n## 4. ESOA\n## 8. ROI\ndeducible fiscal.\n",
        encoding="utf-8")
    ctx = chk._cargar(worktree)
    assert not chk.fuentes_citadas(ctx).passed


def test_fuentes_acepta_con_grafo(worktree):
    ctx = chk._cargar(worktree)   # base cita grafo/LISR
    assert chk.fuentes_citadas(ctx).passed


# --- sin_secretos -----------------------------------------------------------

def test_secretos_detecta_token_expuesto(worktree):
    (worktree / "reporte.md").write_text(
        REPORTE_BASE + "\napi_key = 'sk-abcd1234abcd1234abcd'\n", encoding="utf-8")
    ctx = chk._cargar(worktree)
    assert not chk.sin_secretos(ctx).passed


# --- agregado ---------------------------------------------------------------

def test_correr_todos_devuelve_un_resultado_por_gate(worktree):
    ok, resultados = chk.correr_todos(worktree)
    assert ok
    assert len(resultados) == len(chk.GATES_ACTIVOS)


# --- integracion con el MOTOR del Supervisor (reglas/procesos.toml real) -----
# La ruta que el CLI no ejercita: cargar_configs + correr_gates sobre los
# adaptadores registrados en gates.CHEQUEOS (si borras el registro o el TOML,
# estos tests se ponen rojos).

import subprocess

import gates as gates_mod

DIR_REGLAS = Path(__file__).resolve().parent.parent / "reglas"


def _git_init(worktree: Path):
    # El motor lista archivos cambiados via git (add -A + diff --cached); un
    # worktree real siempre es repo. Sin identidad ni commit alcanza.
    subprocess.run(["git", "init", "-q"], cwd=worktree, check=True)


def test_motor_aprueba_el_paquete_valido_con_el_toml_real(worktree):
    _git_init(worktree)
    configs = gates_mod.cargar_configs(DIR_REGLAS)
    resultados = gates_mod.correr_gates(configs["procesos"], worktree)
    assert len(resultados) == 12  # 12 activos; los 2 de modelo estan inactivos
    fallidos = [(r.regla, r.evidencia) for r in resultados if r.estado != "paso"]
    assert not fallidos, fallidos


def test_motor_traduce_no_ejecutable_cuando_falta_linea_base(worktree):
    _git_init(worktree)
    d = copy.deepcopy(DIAG_BASE); d.pop("linea_base")
    _reescribe(worktree, "diagnostico.yaml", d)
    configs = gates_mod.cargar_configs(DIR_REGLAS)
    [gate_lb] = [g for g in configs["procesos"] if g.regla == "linea_base_cuantificada"]
    [r] = gates_mod.correr_gates([gate_lb], worktree)
    assert r.estado == "no_ejecutable" and r.hallazgos
