#!/usr/bin/env python3
"""Runners de los gates deterministas del departamento de Procesos.

Cada gate es una función `nombre(ctx) -> Chequeo(passed, hallazgo)`. Son
BINARIOS y DETERMINISTAS: no llaman a ningún modelo. Parsean los artefactos que
el Ejecutor deja en el worktree (diagnostico.yaml, reporte.md, build-spec.yaml,
presupuesto.xlsx) y verifican invariantes de estructura/política.

Principio heredado del Supervisor: un gate que no se puede correr (p. ej. falta
una dependencia o un artefacto) NO se asume aprobado — devuelve `no_ejecutable`,
que el Supervisor trata como RECHAZO con hallazgo.

Uso como CLI (para dev/CI):
    python chequeos_procesos.py <ruta-al-worktree>
Sale 0 si todos los gates activos pasan; 1 si alguno falla o no es ejecutable.

Integración: el supervisor-a2a resuelve `chequeos_procesos:<gate>` (ver
reglas/procesos.toml). ⚠️ El Dockerfile del supervisor DEBE copiar este archivo
(como pasó con chequeos_adquisicion.py) o el servicio entra en crash-loop.
"""
from __future__ import annotations

import re
import sys
from dataclasses import dataclass
from pathlib import Path

try:
    import yaml  # PyYAML
except ImportError:  # pragma: no cover
    yaml = None

VEREDICTOS_ESOA = {"eliminar", "simplificar", "optimizar", "automatizar"}
CINCO_S = ["seiri_clasificar", "seiton_ordenar", "seiso_limpiar",
           "seiketsu_estandarizar", "shitsuke_disciplina"]
ALCANCES = {"chico", "mediano", "grande"}

# Marcador de marca blanca: [MAYUSCULAS_O_GUION_BAJO], 2+ chars.
RE_MARCADOR = re.compile(r"\[[A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ_]{1,}\]")
# Palabras que delatan afirmación fiscal/regulatoria (exigen fuente).
RE_FISCAL = re.compile(
    r"\b(deducib\w*|fiscal\w*|regulator\w*|contrat\w*|SAT|LISR|CFF|NIF|"
    r"impuest\w*|permitid\w*|cumplimient\w*)\b", re.IGNORECASE)
# Señal de que SÍ hay fuente/derivación válida.
RE_FUENTE = re.compile(
    r"\b(grafo|fuente|LISR|CFF|NIF|art\.?\s*\d|DOF|ET\s*\d|consultar\s+grafo)\b",
    re.IGNORECASE)
# Secretos comunes (secret-scrubbing heredado).
RE_SECRETO = re.compile(
    r"(sk-[A-Za-z0-9]{16,}|ghp_[A-Za-z0-9]{20,}|service_role|"
    r"eyJ[A-Za-z0-9_\-]{20,}|-----BEGIN [A-Z ]*PRIVATE KEY-----|"
    r"(?:api[_-]?key|password|secret|token)\s*[=:]\s*['\"][^'\"]{8,})",
    re.IGNORECASE)


@dataclass
class Chequeo:
    passed: bool
    hallazgo: str = ""
    no_ejecutable: bool = False


@dataclass
class Contexto:
    """Lo que un gate necesita: el worktree y los artefactos ya cargados."""
    worktree: Path
    diagnostico: dict | None
    build_spec: dict | None
    reporte_txt: str
    raw_texts: dict  # nombre_archivo -> texto crudo (para regex)


def _cargar(worktree: str | Path) -> Contexto:
    wt = Path(worktree)
    raw = {}
    for nombre in ("diagnostico.yaml", "build-spec.yaml", "reporte.md"):
        p = wt / nombre
        raw[nombre] = p.read_text(encoding="utf-8") if p.exists() else ""
    diag = _yaml(raw["diagnostico.yaml"])
    spec = _yaml(raw["build-spec.yaml"])
    return Contexto(wt, diag, spec, raw["reporte.md"], raw)


def _yaml(text: str):
    if not text or yaml is None:
        return None
    try:
        return yaml.safe_load(text)
    except Exception:
        return None


# --- Gates -----------------------------------------------------------------

def estructura_diagnostico(ctx: Contexto) -> Chequeo:
    faltan = [n for n in ("diagnostico.yaml", "reporte.md",
                          "presupuesto.xlsx", "build-spec.yaml")
              if not (ctx.worktree / n).exists()]
    if faltan:
        return Chequeo(False, f"Faltan artefactos: {', '.join(faltan)}")
    if yaml is None:
        return Chequeo(False, "PyYAML no disponible; no se puede validar.",
                       no_ejecutable=True)
    if not isinstance(ctx.diagnostico, dict):
        return Chequeo(False, "diagnostico.yaml no parsea a objeto.")
    req = ["proyecto", "alcance", "pasos_as_is", "cinco_s", "diseno_a2a"]
    faltan_k = [k for k in req if k not in ctx.diagnostico]
    if faltan_k:
        return Chequeo(False, f"diagnostico.yaml sin llaves: {', '.join(faltan_k)}")
    if ctx.diagnostico.get("alcance") not in ALCANCES:
        return Chequeo(False, "alcance debe ser chico|mediano|grande.")
    # Secciones mínimas del reporte legible.
    for marca in ("Resumen ejecutivo", "ESOA", "ROI"):
        if marca.lower() not in ctx.reporte_txt.lower():
            return Chequeo(False, f"reporte.md sin la sección '{marca}'.")
    return Chequeo(True)


def esoa_completo(ctx: Contexto) -> Chequeo:
    d = ctx.diagnostico
    if not isinstance(d, dict) or not isinstance(d.get("pasos_as_is"), list):
        return Chequeo(False, "pasos_as_is ausente o no es lista.",
                       no_ejecutable=True)
    if not d["pasos_as_is"]:
        return Chequeo(False, "pasos_as_is vacío: no hay proceso mapeado.")
    for paso in d["pasos_as_is"]:
        pid = paso.get("id", "?")
        v = str(paso.get("veredicto_esoa", "")).lower()
        if v not in VEREDICTOS_ESOA:
            return Chequeo(False, f"Paso {pid}: veredicto_esoa inválido ('{v}').")
        if not str(paso.get("justificacion", "")).strip():
            return Chequeo(False, f"Paso {pid}: justificacion vacía.")
    return Chequeo(True)


def cinco_s_aplicado(ctx: Contexto) -> Chequeo:
    d = ctx.diagnostico
    cinco = d.get("cinco_s") if isinstance(d, dict) else None
    if not isinstance(cinco, dict):
        return Chequeo(False, "cinco_s ausente o no es objeto.",
                       no_ejecutable=True)
    for s in CINCO_S:
        if not str(cinco.get(s, "")).strip():
            return Chequeo(False, f"5S sin evaluar: '{s}' vacío "
                                  f"(usa un hallazgo o 'n/a' con razón).")
    return Chequeo(True)


def control_humano_por_automatizacion(ctx: Contexto) -> Chequeo:
    d = ctx.diagnostico
    disenos = d.get("diseno_a2a") if isinstance(d, dict) else None
    if not isinstance(disenos, list):
        return Chequeo(False, "diseno_a2a ausente o no es lista.",
                       no_ejecutable=True)
    for item in disenos:
        nombre = item.get("automatizacion", "?")
        ch = str(item.get("control_humano", "")).strip()
        if not ch:
            return Chequeo(False, f"Automatización '{nombre}' sin control_humano.")
        if re.search(r"cero\s+human|sin\s+human|ningun\s+human", ch, re.I):
            return Chequeo(False, f"Automatización '{nombre}': 'cero humanos' "
                                  f"no permitido.")
    return Chequeo(True)


def presupuesto_dos_monedas(ctx: Contexto) -> Chequeo:
    p = ctx.worktree / "presupuesto.xlsx"
    if not p.exists():
        return Chequeo(False, "Falta presupuesto.xlsx.")
    try:
        from openpyxl import load_workbook
    except ImportError:
        return Chequeo(False, "openpyxl no disponible; no se puede validar.",
                       no_ejecutable=True)
    try:
        wb = load_workbook(p, read_only=True, data_only=True)
    except Exception as e:
        return Chequeo(False, f"presupuesto.xlsx ilegible: {e}",
                       no_ejecutable=True)
    texto = " ".join(
        str(c) for ws in wb.worksheets for row in ws.iter_rows(values_only=True)
        for c in row if c is not None)
    tiene_usd = "USD" in texto
    tiene_mxn = "MXN" in texto
    tiene_tc = re.search(r"tipo de cambio", texto, re.I) or "TC" in texto
    if not (tiene_usd and tiene_mxn):
        return Chequeo(False, "El presupuesto no muestra MXN y USD.")
    if not tiene_tc:
        return Chequeo(False, "El presupuesto no declara el tipo de cambio (supuesto).")
    return Chequeo(True)


def build_spec_valida(ctx: Contexto) -> Chequeo:
    s = ctx.build_spec
    if not isinstance(s, dict):
        return Chequeo(False, "build-spec.yaml ausente o no parsea.",
                       no_ejecutable=True)
    items = s.get("construir")
    if not isinstance(items, list) or not items:
        return Chequeo(False, "build-spec sin lista 'construir'.")
    req = ["departamento_destino", "skills_requeridas", "clis_requeridos",
           "control_humano", "gate_humano_irreversible"]
    for it in items:
        iid = it.get("id", it.get("automatizacion", "?"))
        for k in req:
            if k not in it:
                return Chequeo(False, f"build-spec item '{iid}' sin '{k}'.")
        if not isinstance(it["skills_requeridas"], list) or \
           not isinstance(it["clis_requeridos"], list):
            return Chequeo(False, f"item '{iid}': skills/clis deben ser listas.")
        if not isinstance(it["gate_humano_irreversible"], bool):
            return Chequeo(False, f"item '{iid}': gate_humano_irreversible debe ser bool.")
    disparo = s.get("disparo", {})
    if disparo.get("requiere_aprobacion_humana") is not True:
        return Chequeo(False, "disparo.requiere_aprobacion_humana debe ser true "
                              "(candado: no se construye sin OK humano).")
    return Chequeo(True)


def sin_marcadores(ctx: Contexto) -> Chequeo:
    for nombre, txt in ctx.raw_texts.items():
        m = RE_MARCADOR.search(txt)
        if m:
            return Chequeo(False, f"{nombre}: marcador de marca blanca sin "
                                  f"sustituir: {m.group(0)}")
    return Chequeo(True)


def fuentes_citadas(ctx: Contexto) -> Chequeo:
    # Si el diagnóstico afirma algo fiscal/regulatorio, debe haber fuente/grafo.
    txt = ctx.raw_texts.get("diagnostico.yaml", "") + "\n" + ctx.reporte_txt
    if RE_FISCAL.search(txt) and not RE_FUENTE.search(txt):
        return Chequeo(False, "Hay afirmaciones fiscales/regulatorias sin fuente "
                              "ni marca 'consultar grafo'.")
    return Chequeo(True)


def sin_secretos(ctx: Contexto) -> Chequeo:
    for nombre, txt in ctx.raw_texts.items():
        m = RE_SECRETO.search(txt)
        if m:
            return Chequeo(False, f"{nombre}: posible secreto expuesto.")
    return Chequeo(True)


def linea_base_cuantificada(ctx: Contexto) -> Chequeo:
    """La línea base (cuánto vale el proceso hoy) es el ancla del ROI."""
    d = ctx.diagnostico
    lb = d.get("linea_base") if isinstance(d, dict) else None
    if not isinstance(lb, dict):
        return Chequeo(False, "Falta linea_base (cuánto vale el proceso hoy).",
                       no_ejecutable=True)
    mensual = lb.get("costo_actual_mensual_usd")
    anual = lb.get("costo_actual_anual_usd")
    tiene_costo = any(isinstance(v, (int, float)) and v > 0
                      for v in (mensual, anual))
    if not tiene_costo:
        return Chequeo(False, "linea_base sin costo actual (mensual o anual > 0).")
    # Si es estimada, debe declarar supuestos (no hay diagnóstico sin datos).
    if lb.get("es_estimado") and not lb.get("supuestos"):
        return Chequeo(False, "linea_base estimada sin 'supuestos' declarados.")
    return Chequeo(True)


def consejo_y_reto(ctx: Contexto) -> Chequeo:
    """Recomendación clara + pase adversarial honesto; el reto no va vacío."""
    d = ctx.diagnostico
    if not isinstance(d, dict):
        return Chequeo(False, "diagnostico.yaml no parsea.", no_ejecutable=True)
    if not str(d.get("consejo", "")).strip():
        return Chequeo(False, "Falta 'consejo' (la recomendación).")
    retos = d.get("reto_limitantes")
    if not isinstance(retos, list) or not any(str(r).strip() for r in retos):
        return Chequeo(False, "Falta 'reto_limitantes' o va vacío "
                              "(un diagnóstico sin límites no se revisó a sí mismo).")
    return Chequeo(True)


def herramientas_en_stack(ctx: Contexto) -> Chequeo:
    """No se proponen herramientas fuera del stack del cliente sin justificar."""
    s = ctx.build_spec
    if not isinstance(s, dict):
        return Chequeo(False, "build-spec.yaml no parsea.", no_ejecutable=True)
    items = s.get("construir") or []
    propone = any(it.get("herramientas_propuestas") for it in items)
    stack = s.get("stack_cliente")
    if not isinstance(stack, dict) or not stack.get("herramientas"):
        if propone:
            return Chequeo(False, "Se proponen herramientas sin declarar "
                                  "stack_cliente (pregunta el stack antes de proponer).")
        return Chequeo(True)
    en_stack = {str(h).lower() for h in stack.get("herramientas", [])}
    for it in items:
        iid = it.get("id", it.get("automatizacion", "?"))
        just = str(it.get("justificacion_herramientas", "")).strip()
        for h in it.get("herramientas_propuestas") or []:
            if str(h).lower() not in en_stack and not just:
                return Chequeo(False, f"item '{iid}': herramienta '{h}' fuera del "
                                      f"stack sin justificacion_herramientas.")
    return Chequeo(True)


GATES_ACTIVOS = [
    ("estructura_diagnostico", estructura_diagnostico),
    ("linea_base_cuantificada", linea_base_cuantificada),
    ("esoa_completo", esoa_completo),
    ("cinco_s_aplicado", cinco_s_aplicado),
    ("control_humano_por_automatizacion", control_humano_por_automatizacion),
    ("consejo_y_reto", consejo_y_reto),
    ("presupuesto_dos_monedas", presupuesto_dos_monedas),
    ("build_spec_valida", build_spec_valida),
    ("herramientas_en_stack", herramientas_en_stack),
    ("sin_marcadores", sin_marcadores),
    ("fuentes_citadas", fuentes_citadas),
    ("sin_secretos", sin_secretos),
]


def correr_todos(worktree: str | Path) -> tuple[bool, list]:
    ctx = _cargar(worktree)
    resultados = []
    ok = True
    for nombre, fn in GATES_ACTIVOS:
        ch = fn(ctx)
        resultados.append((nombre, ch))
        if not ch.passed:
            ok = False
    return ok, resultados


# ---------- registro en el motor del Supervisor (gates.CHEQUEOS) ----------
# Import unidireccional, mismo patron que chequeos_adquisicion/chequeos_fabric:
# este modulo importa de `gates` y SE REGISTRA al importarse (executor.py lo
# importa); gates.py no lo conoce. Los gates de procesos operan sobre los
# ARTEFACTOS del paquete to-be (diagnostico.yaml, build-spec.yaml, reporte.md,
# presupuesto.xlsx), no sobre la lista de archivos cambiados: el adaptador
# carga el Contexto del worktree y traduce Chequeo -> ResultadoGate.
# `sin_secretos` NO se registra: reglas/procesos.toml reusa el chequeo base de
# gates.py (mismo criterio compartido que adquisicion); la version de este
# modulo queda solo para el modo CLI.
import gates
from gates import ResultadoGate


def _adaptar(nombre, fn):
    def chequeo(gate, worktree, archivos):
        ch = fn(_cargar(worktree))
        if ch.passed:
            return ResultadoGate(
                regla=gate.regla, estado="paso",
                evidencia=f"{nombre}: paquete to-be sin hallazgos",
            )
        estado = "no_ejecutable" if ch.no_ejecutable else "fallo"
        return ResultadoGate(
            regla=gate.regla, estado=estado, evidencia=ch.hallazgo,
            hallazgos=[{"regla": gate.regla, "evidencia": ch.hallazgo}],
        )
    return chequeo


CHEQUEOS_PROCESOS = {
    nombre: _adaptar(nombre, fn)
    for nombre, fn in GATES_ACTIVOS
    if nombre != "sin_secretos"
}
gates.CHEQUEOS.update(CHEQUEOS_PROCESOS)


def main(argv):
    if len(argv) != 2:
        print("uso: python chequeos_procesos.py <ruta-al-worktree>")
        return 2
    ok, resultados = correr_todos(argv[1])
    for nombre, ch in resultados:
        if ch.passed:
            estado = "PASS"
        elif ch.no_ejecutable:
            estado = "NO_EJECUTABLE(=rechazo)"
        else:
            estado = "FAIL"
        linea = f"[{estado}] {nombre}"
        if ch.hallazgo:
            linea += f" — {ch.hallazgo}"
        print(linea)
    print("\nVEREDICTO:", "APROBADO" if ok else "RECHAZADO")
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main(sys.argv))
