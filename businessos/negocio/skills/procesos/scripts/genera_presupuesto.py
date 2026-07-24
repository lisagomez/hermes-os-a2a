#!/usr/bin/env python3
"""Genera el presupuesto de un proyecto de diagnóstico + implementación A2A.

Produce un Excel (.xlsx) editable con las hojas: Supuestos, Esfuerzo,
Presupuesto (MXN y USD), Tiempo humano-agente y ROI. Las cifras por defecto
coinciden con references/costeo-pricing.md. Todos los valores son de
referencia y configurables: valídalos antes de cotizar a un cliente.

Uso básico:
    python genera_presupuesto.py --alcance mediano --tc 18.5 --margen 0.35 \
        --salida presupuesto_cliente.xlsx

Alcance a la medida: pasa --config archivo.json (ver --help y el README del
skill). Los parámetros de ROI también se pueden sobreescribir por CLI.
"""
import argparse
import json
import sys

# --- Datos de referencia (deben coincidir con costeo-pricing.md) ---------

# Tarifa de COSTO interno por rol (USD/hora).
RATES = {
    "Consultor": 90.0,     # Consultor líder / Arquitecto de procesos IA
    "IngAgentes": 70.0,    # Ingeniero de agentes de IA
    "Integracion": 65.0,   # Especialista en integración (APIs/datos)
    "QA": 45.0,            # QA / Validación
    "PM": 55.0,           # Gestión de proyecto
    "Change": 50.0,       # Change management / Adopción
}
ROLE_LABEL = {
    "Consultor": "Consultor líder / Arquitecto IA",
    "IngAgentes": "Ingeniero de agentes de IA",
    "Integracion": "Especialista en integración",
    "QA": "QA / Validación",
    "PM": "Gestión de proyecto (PM)",
    "Change": "Change management / Adopción",
}

TIERS = {"chico": 0, "mediano": 1, "grande": 2}

# (Fase, rol, [horas_chico, horas_mediano, horas_grande])
PHASES = [
    ("1. Descubrimiento y diagnóstico (ESOA)", "Consultor", [16, 32, 60]),
    ("2. Diseño de solución A2A", "Consultor", [12, 28, 55]),
    ("3. Implementación (build de agentes)", "IngAgentes", [24, 70, 160]),
    ("4. Integración de sistemas y datos", "Integracion", [12, 40, 110]),
    ("5. Pruebas y validación", "QA", [10, 28, 65]),
    ("6. Despliegue y adopción", "Change", [8, 20, 45]),
    ("7. Soporte / hypercare", "IngAgentes", [8, 16, 40]),
]
TOOLS = [500.0, 1500.0, 4000.0]          # herramientas/infra por alcance (USD)
PM_PCT = 0.15
CONTINGENCY = 0.10

# Parámetros de ROI por alcance: N corridas/mes, H_h, H_s, costo_hora_op (USD)
ROI_DEFAULTS = {
    "chico":   {"N": 60,  "H_h": 1.0, "H_s": 0.15, "costo_hora_op": 20.0},
    "mediano": {"N": 200, "H_h": 1.5, "H_s": 0.20, "costo_hora_op": 25.0},
    "grande":  {"N": 500, "H_h": 2.0, "H_s": 0.25, "costo_hora_op": 30.0},
}


def build_model(alcance, tc, margen, rates, phases, tools, roi):
    idx = TIERS[alcance]
    rows = []
    horas_totales = 0.0
    mano_obra = 0.0
    for name, role, hrs in phases:
        h = hrs[idx] if isinstance(hrs, list) else hrs
        rate = rates[role]
        costo = h * rate
        rows.append({"fase": name, "rol": ROLE_LABEL[role],
                     "horas": h, "tarifa_costo": rate, "costo_usd": costo})
        horas_totales += h
        mano_obra += costo
    pm_horas = round(PM_PCT * horas_totales, 1)
    pm_costo = pm_horas * rates["PM"]
    rows.append({"fase": "PM (15% transversal)", "rol": ROLE_LABEL["PM"],
                 "horas": pm_horas, "tarifa_costo": rates["PM"],
                 "costo_usd": pm_costo})
    mano_obra += pm_costo
    horas_totales += pm_horas

    subtotal = mano_obra + tools
    conting = CONTINGENCY * subtotal
    costo_total = subtotal + conting
    precio = costo_total * (1 + margen)

    # Línea base: cuánto vale/cuesta el proceso HOY (antes del rediseño).
    # Es el "antes" contra el que se mide todo el ROI.
    N, H_h, H_s = roi["N"], roi["H_h"], roi["H_s"]
    cop = roi["costo_hora_op"]
    base_horas_mes = N * H_h
    base_mensual = base_horas_mes * cop
    base_anual = base_mensual * 12

    # ROI (el "después" contra la línea base)
    ahorro_horas_mes = N * (H_h - H_s)
    ahorro_mes = ahorro_horas_mes * cop
    ahorro_anio = ahorro_mes * 12
    pct_agente = (H_h - H_s) / H_h if H_h else 0
    pct_humano = H_s / H_h if H_h else 0
    payback = precio / ahorro_mes if ahorro_mes else float("inf")
    roi_anio1 = (ahorro_anio - precio) / precio if precio else 0

    return {
        "alcance": alcance, "tc": tc, "margen": margen,
        "rows": rows, "horas_totales": horas_totales,
        "mano_obra": mano_obra, "tools": tools, "subtotal": subtotal,
        "contingencia": conting, "costo_total": costo_total, "precio": precio,
        "linea_base": {
            "horas_mes": base_horas_mes, "mensual": base_mensual,
            "anual": base_anual,
        },
        "roi": {
            "N": N, "H_h": H_h, "H_s": H_s, "costo_hora_op": cop,
            "ahorro_horas_mes": ahorro_horas_mes, "ahorro_mes": ahorro_mes,
            "ahorro_anio": ahorro_anio, "pct_agente": pct_agente,
            "pct_humano": pct_humano, "payback": payback, "roi_anio1": roi_anio1,
        },
    }


def write_xlsx(m, salida):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        sys.exit("Falta openpyxl. Instala con: pip install openpyxl "
                 "--break-system-packages")

    tc = m["tc"]
    wb = Workbook()
    hdr_fill = PatternFill("solid", fgColor="1F2937")
    hdr_font = Font(color="FFFFFF", bold=True)
    accent = PatternFill("solid", fgColor="BEF264")
    bold = Font(bold=True)
    money = '#,##0'
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(ws, row, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    # --- Supuestos ---
    ws = wb.active
    ws.title = "Supuestos"
    ws["A1"] = "Presupuesto A2A — Supuestos"
    ws["A1"].font = Font(bold=True, size=14)
    data = [
        ("Alcance", m["alcance"]),
        ("Tipo de cambio (MXN por USD)", tc),
        ("Margen sobre costo", m["margen"]),
        ("PM (% de horas de fases)", PM_PCT),
        ("Contingencia", CONTINGENCY),
        ("Corridas/mes (N)", m["roi"]["N"]),
        ("Horas-humano por corrida hoy (H_h)", m["roi"]["H_h"]),
        ("Horas-humano supervisión después (H_s)", m["roi"]["H_s"]),
        ("Costo por hora operativo del cliente (USD)", m["roi"]["costo_hora_op"]),
    ]
    r = 3
    for k, v in data:
        ws.cell(row=r, column=1, value=k).font = bold
        ws.cell(row=r, column=2, value=v)
        r += 1
    ws.cell(row=r + 1, column=1,
            value="Valores de referencia y configurables. Valídalos antes de "
                  "cotizar. Confirma el tipo de cambio el día de la cotización.")
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 24

    # --- Esfuerzo por fase ---
    ws = wb.create_sheet("Esfuerzo")
    headers = ["Fase", "Rol", "Horas", "Tarifa costo (USD/h)", "Costo (USD)"]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for row in m["rows"]:
        ws.append([row["fase"], row["rol"], row["horas"],
                   row["tarifa_costo"], round(row["costo_usd"])])
    tot_r = ws.max_row + 1
    ws.cell(row=tot_r, column=1, value="TOTAL").font = bold
    ws.cell(row=tot_r, column=3, value=m["horas_totales"]).font = bold
    ws.cell(row=tot_r, column=5, value=round(m["mano_obra"])).font = bold
    for row in ws.iter_rows(min_row=2, max_row=tot_r):
        row[4].number_format = money
    for col, w in zip("ABCDE", (40, 32, 10, 20, 16)):
        ws.column_dimensions[col].width = w

    # --- Presupuesto ---
    ws = wb.create_sheet("Presupuesto")
    ws.append(["Concepto", "USD", "MXN"])
    style_header(ws, 1, 3)
    lines = [
        ("Mano de obra (fases + PM)", m["mano_obra"]),
        ("Herramientas / infraestructura", m["tools"]),
        ("Subtotal", m["subtotal"]),
        ("Contingencia (10%)", m["contingencia"]),
        ("COSTO TOTAL", m["costo_total"]),
        (f"PRECIO (margen {int(m['margen']*100)}%)", m["precio"]),
    ]
    for label, usd in lines:
        ws.append([label, round(usd), round(usd * tc)])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[1].number_format = money
        row[2].number_format = money
    # resaltar COSTO y PRECIO
    for rr in (ws.max_row - 1, ws.max_row):
        for c in range(1, 4):
            ws.cell(row=rr, column=c).font = bold
    for c in range(1, 4):
        ws.cell(row=ws.max_row, column=c).fill = accent
    for col, w in zip("ABC", (34, 16, 16)):
        ws.column_dimensions[col].width = w

    # --- Tiempo humano-agente ---
    ws = wb.create_sheet("Humano-Agente")
    ws["A1"] = "Reparto de tiempo humano vs. agente (operación)"
    ws["A1"].font = Font(bold=True, size=12)
    roi = m["roi"]
    rows = [
        ("Corridas por mes", roi["N"]),
        ("Horas-humano por corrida — hoy (as-is)", roi["H_h"]),
        ("Horas-humano supervisión por corrida — después", roi["H_s"]),
        ("Horas-humano/mes hoy", round(roi["N"] * roi["H_h"], 1)),
        ("Horas-humano/mes después (supervisión)", round(roi["N"] * roi["H_s"], 1)),
        ("Horas-humano liberadas/mes", round(roi["ahorro_horas_mes"], 1)),
        ("% del trabajo que ejecutan los agentes", f"{roi['pct_agente']*100:.0f}%"),
        ("% que queda como supervisión humana", f"{roi['pct_humano']*100:.0f}%"),
    ]
    r = 3
    for k, v in rows:
        ws.cell(row=r, column=1, value=k).font = bold
        ws.cell(row=r, column=2, value=v)
        r += 1
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 16

    # --- Línea base + ROI ---
    ws = wb.create_sheet("Linea-base-ROI")
    ws.append(["Métrica", "USD", "MXN"])
    style_header(ws, 1, 3)
    base = m["linea_base"]
    roi_lines = [
        ("LÍNEA BASE — costo actual del proceso / mes", base["mensual"]),
        ("LÍNEA BASE — costo actual del proceso / año", base["anual"]),
        ("Ahorro mensual", roi["ahorro_mes"]),
        ("Ahorro anual", roi["ahorro_anio"]),
        ("Inversión (precio del proyecto)", m["precio"]),
    ]
    for label, usd in roi_lines:
        ws.append([label, round(usd), round(usd * tc)])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[1].number_format = money
        row[2].number_format = money
    for c in range(1, 4):                     # resaltar las 2 filas de línea base
        ws.cell(row=2, column=c).font = bold
        ws.cell(row=3, column=c).font = bold
        ws.cell(row=2, column=c).fill = accent
    ws.append([])
    pr = ws.max_row + 1
    ws.cell(row=pr, column=1, value="Payback (meses)").font = bold
    ws.cell(row=pr, column=2,
            value=round(roi["payback"], 1) if roi["payback"] != float("inf") else "n/d")
    ws.cell(row=pr + 1, column=1, value="ROI año 1").font = bold
    ws.cell(row=pr + 1, column=2, value=f"{roi['roi_anio1']*100:.0f}%")
    for col, w in zip("ABC", (30, 16, 16)):
        ws.column_dimensions[col].width = w

    wb.save(salida)
    return salida


def load_config(path, alcance, tc, margen):
    """Permite sobreescribir horas/tarifas/tools/roi con un JSON."""
    with open(path, encoding="utf-8") as f:
        cfg = json.load(f)
    rates = dict(RATES)
    rates.update(cfg.get("rates", {}))
    tools = cfg.get("tools")
    phases = PHASES
    if "phases" in cfg:
        phases = [(p["fase"], p["rol"], p["horas"]) for p in cfg["phases"]]
    roi = dict(ROI_DEFAULTS[alcance])
    roi.update(cfg.get("roi", {}))
    if tools is None:
        tools = TOOLS[TIERS[alcance]]
    return rates, phases, tools, roi


def main():
    ap = argparse.ArgumentParser(description="Genera el presupuesto A2A en xlsx.")
    ap.add_argument("--alcance", choices=list(TIERS), default="mediano")
    ap.add_argument("--tc", type=float, default=18.5,
                    help="Tipo de cambio MXN por USD (confírmalo al cotizar).")
    ap.add_argument("--margen", type=float, default=0.35,
                    help="Margen sobre costo (0.35 = 35%%).")
    ap.add_argument("--salida", default="presupuesto_a2a.xlsx")
    ap.add_argument("--config", help="JSON con overrides (horas, tarifas, roi).")
    # Overrides de ROI por CLI
    ap.add_argument("--corridas", type=int)
    ap.add_argument("--hh", type=float, help="Horas-humano por corrida hoy.")
    ap.add_argument("--hs", type=float, help="Horas supervisión por corrida.")
    ap.add_argument("--costo-hora-op", type=float)
    args = ap.parse_args()

    if args.config:
        rates, phases, tools, roi = load_config(args.config, args.alcance,
                                                args.tc, args.margen)
    else:
        rates, phases = RATES, PHASES
        tools = TOOLS[TIERS[args.alcance]]
        roi = dict(ROI_DEFAULTS[args.alcance])

    if args.corridas is not None:
        roi["N"] = args.corridas
    if args.hh is not None:
        roi["H_h"] = args.hh
    if args.hs is not None:
        roi["H_s"] = args.hs
    if args.costo_hora_op is not None:
        roi["costo_hora_op"] = args.costo_hora_op

    m = build_model(args.alcance, args.tc, args.margen, rates, phases, tools, roi)
    out = write_xlsx(m, args.salida)

    print(f"Presupuesto generado: {out}")
    print(f"  Alcance: {m['alcance']} | TC: {m['tc']} | margen: {m['margen']}")
    print(f"  Costo total: USD {m['costo_total']:,.0f} | "
          f"MXN {m['costo_total']*m['tc']:,.0f}")
    print(f"  Línea base:  USD {m['linea_base']['mensual']:,.0f}/mes | "
          f"USD {m['linea_base']['anual']:,.0f}/año (costo actual del proceso)")
    print(f"  PRECIO:      USD {m['precio']:,.0f} | "
          f"MXN {m['precio']*m['tc']:,.0f}")
    print(f"  Ahorro/mes:  USD {m['roi']['ahorro_mes']:,.0f} | "
          f"payback {m['roi']['payback']:.1f} meses | "
          f"ROI año1 {m['roi']['roi_anio1']*100:.0f}%")
    print(f"  Reparto: {m['roi']['pct_agente']*100:.0f}% agentes / "
          f"{m['roi']['pct_humano']*100:.0f}% supervisión humana")


if __name__ == "__main__":
    main()
